"""
========================================================================
  KRAFT HEINZ (NASDAQ: KHC) - MODELO FINANCIERO PROYECTADO 5 AÑOS
  Prueba Técnica - NoNighter LLC  |  Posicion: Modelador Financiero
------------------------------------------------------------------------
  Fuente unica de cifras historicas: 10-K FY2025 (fiscal year ended
  December 27, 2025), presentado por The Kraft Heinz Company ante la SEC.

  Salida: KHC_Financial_Model.xlsx
  Requiere: openpyxl  (pip install openpyxl)
========================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ============================================================
# 1) CONSTANTES DE FORMATO (segun consigna)
# ============================================================
FONT_NAME       = "Arial"
FONT_SIZE       = 8
COLOR_HARDCODE  = "0000FF"   # azul - inputs/hardcodes
COLOR_LINK      = "3C7D22"   # verde - links entre sheets
COLOR_FORMULA   = "000000"   # negro - formulas dentro de la misma hoja
COLOR_HEADER_BG = "1F4E78"   # azul oscuro - fondo de cabeceras
COLOR_HEADER_FG = "FFFFFF"   # blanco - texto de cabeceras
COLOR_SUBTOTAL  = "D9E1F2"   # gris-azul claro - subtotales
COLOR_CHECK_OK  = "C6EFCE"   # verde claro - check OK
COLOR_CHECK_ERR = "FFC7CE"   # rojo claro - check error

NUM_FMT      = '#,##0;(#,##0);"-"'
NUM_FMT_USD  = '$#,##0;($#,##0);"-"'
PCT_FMT      = '0.0%;(0.0%);"-"'
DAYS_FMT     = '0.0" d"'
MULT_FMT     = '0.00x'

THIN = Side(border_style="thin", color="BFBFBF")
TOP_BORDER    = Border(top=THIN)
BOT_BORDER    = Border(bottom=THIN)
TOP_BOT       = Border(top=THIN, bottom=THIN)
DBL_BOT       = Border(bottom=Side(border_style="double", color="000000"))

# ============================================================
# 2) ESTRUCTURA DE PERIODOS
# ============================================================
# Columnas:  A = labels  |  B..D = historicos  |  E..I = forecast
# B=FY2023A  C=FY2024A  D=FY2025A  E=FY2026E  F=FY2027E  G=FY2028E  H=FY2029E  I=FY2030E
HIST_LABELS = ["FY2023A", "FY2024A", "FY2025A"]
FCST_LABELS = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
ALL_LABELS  = HIST_LABELS + FCST_LABELS
N_HIST = len(HIST_LABELS)         # 3
N_FCST = len(FCST_LABELS)         # 5
N_ALL  = N_HIST + N_FCST          # 8

FIRST_DATA_COL = 2                # columna B
HIST_COLS = [get_column_letter(FIRST_DATA_COL + i) for i in range(N_HIST)]          # ['B','C','D']
FCST_COLS = [get_column_letter(FIRST_DATA_COL + N_HIST + i) for i in range(N_FCST)] # ['E','F','G','H','I']
ALL_COLS  = HIST_COLS + FCST_COLS                                                   # ['B'..'I']
LAST_HIST_COL = HIST_COLS[-1]    # 'D' (FY25)
FIRST_FCST_COL = FCST_COLS[0]    # 'E' (FY26)

# ============================================================
# 3) DATOS HISTORICOS DEL 10-K FY2025 (en $M)
# ============================================================
# --- Net Sales por segmento (Note 21, Segment Reporting) ---
NA_SALES_HIST  = [20126, 19543, 18586]
IDM_SALES_HIST = [3623,  3535,  3539]
EM_SALES_HIST  = [2891,  2768,  2817]

# --- Adjusted COGS por segmento (Note 21) ---
NA_COGS_HIST   = [12948, 12356, 12076]
IDM_COGS_HIST  = [2580,  2482,  2497]
# EM no se reporta directamente; lo derivamos del total reportado
TOTAL_COGS_HIST = [17714, 16878, 16633]   # P&L consolidado
EM_COGS_HIST = [TOTAL_COGS_HIST[i] - NA_COGS_HIST[i] - IDM_COGS_HIST[i] for i in range(3)]

# --- SG&A excluyendo impairments (P&L consolidado) ---
SGA_EX_IMPAIR_HIST = [3692, 3616, 3672]

# --- D&A por segmento (Note 21) ---
NA_DA_HIST   = [561, 614, 638]
IDM_DA_HIST  = [157, 156, 150]
EM_DA_HIST   = [157, 106, 112]
CORP_DA_HIST = [86,   72,  68]
TOT_DA_HIST  = [961, 948, 968]

# --- Depreciation expense (Note 8) ---
DEP_HIST = [710, 696, 722]
# Amortizacion implicita = D&A total - Depreciation
AMORT_HIST = [TOT_DA_HIST[i] - DEP_HIST[i] for i in range(3)]  # ~[251, 252, 246]

# --- CapEx por segmento (Note 21) ---
NA_CAPEX_HIST   = [604, 643, 497]
IDM_CAPEX_HIST  = [178, 154, 112]
EM_CAPEX_HIST   = [163, 115, 105]
CORP_CAPEX_HIST = [68,  112,  87]
TOT_CAPEX_HIST  = [1013, 1024, 801]

# --- P&L consolidado (otros items) ---
GW_IMPAIR_HIST   = [510,  1638, 6734]   # goodwill impairments
INT_IMPAIR_HIST  = [152,  2031, 2572]   # intangible impairments
OP_INC_HIST      = [4572, 1683, -4669]  # operating income/(loss)
INT_EXP_HIST     = [912,  912,   947]
OTHER_EXP_HIST   = [27,  -85,   -171]   # other expense/(income)
PRETAX_HIST      = [3633, 856,  -5445]
TAX_PROV_HIST    = [787, -1890,  403]
NI_HIST          = [2846, 2746, -5848]

# --- Balance Sheet (10-K solo presenta FY24 y FY25) ---
# Activos corrientes
CASH_24, CASH_25       = 1334, 2615
AR_24, AR_25           = 2147, 2254
INV_24, INV_25         = 3376, 3167
PREPAID_24, PREPAID_25 = 215,  291
MKTSEC_24, MKTSEC_25   = 0,    1060
OTH_CA_24, OTH_CA_25   = 583,  588
AHFS_24, AHFS_25       = 0,    152

# No corrientes
PPE_24, PPE_25         = 7152, 7318
GW_24, GW_25           = 28673, 22179
INTANG_24, INTANG_25   = 40099, 37529
OTH_NCA_24, OTH_NCA_25 = 4708, 4633

# Pasivos corrientes
CURLTD_24, CURLTD_25   = 654,  1908
AP_24, AP_25           = 4188, 4308
ACCMKT_24, ACCMKT_25   = 697,  801
INTPAY_24, INTPAY_25   = 263,  298
OTH_CL_24, OTH_CL_25   = 1451, 1455
LHFS_24, LHFS_25       = 0,    8

# No corrientes
LTD_24, LTD_25         = 19215, 19311
DEFTAX_24, DEFTAX_25   = 9679, 9022
ACCPOST_24, ACCPOST_25 = 135,  131
LTDEFINC_24, LTDEFINC_25 = 1374, 1321
OTH_NCL_24, OTH_NCL_25 = 1306, 1434

# Equity
CSTK_24, CSTK_25       = 12,   12
APIC_24, APIC_25       = 52135, 51287
RE_24, RE_25           = 2171, -4629
AOCI_24, AOCI_25       = -2915, -2370
TREAS_24, TREAS_25     = -2218, -2636
NCI_24, NCI_25         = 134,  113
REDNCI_24, REDNCI_25   = 6,    12

# --- Debt schedule: principal maturities (Note 17) ---
DEBT_MAT = {2026: 1879, 2027: 1893, 2028: 1679, 2029: 1009, 2030: 915}

# --- CapEx + cash flow items (Statement of Cash Flows) ---
DIV_PAID_HIST    = [1965, 1931, 1898]
SHARE_REPCH_HIST = [455,  988,  436]

# --- Cantidad de acciones diluidas (en M) ---
# 2025 = NI / EPS dil = -5846 / -4.93 = 1186M (aprox); usamos shares outstanding actuales
SHARES_OUT_25 = 1184   # acciones en circulacion (millones)

# ============================================================
# 4) HELPERS DE ESTILO
# ============================================================
def base_font(color=COLOR_FORMULA, bold=False, italic=False):
    return Font(name=FONT_NAME, size=FONT_SIZE, color=color, bold=bold, italic=italic)

def apply_default_font(ws, max_row=200, max_col=15):
    """Aplica Arial 8 a todas las celdas usadas."""
    f = base_font()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font.name is None or cell.font.name != FONT_NAME:
                cell.font = f

def set_cell(ws, coord, value, *, kind="formula", bold=False, italic=False,
             fmt=None, fill=None, border=None, halign=None, comment=None):
    """
    kind: 'hardcode' (azul) | 'link' (verde) | 'formula' (negro) | 'label'
    """
    cell = ws[coord]
    cell.value = value
    color = COLOR_FORMULA
    if kind == "hardcode":
        color = COLOR_HARDCODE
    elif kind == "link":
        color = COLOR_LINK
    cell.font = base_font(color=color, bold=bold, italic=italic)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)
    if border:
        cell.border = border
    if halign:
        cell.alignment = Alignment(horizontal=halign, vertical="center")
    if comment:
        cell.comment = Comment(comment, "Model")
    return cell

def write_row_values(ws, row, values, *, start_col=FIRST_DATA_COL, kind="hardcode", fmt=NUM_FMT):
    """Escribe una lista de valores empezando en (row, start_col)."""
    for i, v in enumerate(values):
        coord = f"{get_column_letter(start_col + i)}{row}"
        set_cell(ws, coord, v, kind=kind, fmt=fmt)

def section_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = base_font(color=COLOR_HEADER_FG, bold=True)
    cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    # Pintar todas las columnas A..I
    for c in range(2, 2 + N_ALL):
        cc = ws.cell(row=row, column=c)
        cc.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cc.font = base_font(color=COLOR_HEADER_FG, bold=True)

def period_headers(ws, row):
    set_cell(ws, f"A{row}", "($M)", bold=True, italic=True, halign="left")
    for i, lbl in enumerate(ALL_LABELS):
        coord = f"{get_column_letter(FIRST_DATA_COL + i)}{row}"
        c = ws[coord]
        c.value = lbl
        c.font = base_font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = BOT_BORDER

def set_label(ws, row, text, *, bold=False, italic=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + text)
    cell.font = base_font(bold=bold, italic=italic)

def column_widths(ws):
    ws.column_dimensions['A'].width = 46
    for c in ALL_COLS:
        ws.column_dimensions[c].width = 11

def freeze_view(ws, row, col_letter="B"):
    ws.freeze_panes = f"{col_letter}{row}"

# ============================================================
# 5) CONSTRUCCION DEL WORKBOOK
# ============================================================
wb = Workbook()
wb.remove(wb.active)   # eliminamos la hoja default

# ------------------------------------------------------------
# HOJA 1 - COVER
# ------------------------------------------------------------
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    set_cell(ws, "B2", "THE KRAFT HEINZ COMPANY (NASDAQ: KHC)", bold=True)
    ws["B2"].font = Font(name=FONT_NAME, size=18, bold=True, color="1F4E78")

    set_cell(ws, "B3", "Modelo Financiero Proyectado - 5 anios forward", bold=True)
    ws["B3"].font = Font(name=FONT_NAME, size=12, italic=True, color="404040")

    set_cell(ws, "B5", "Prueba Tecnica  |  NoNighter LLC  |  Posicion: Modelador Financiero", bold=True)

    set_cell(ws, "B7",  "Fuente unica historica:",  bold=True)
    set_cell(ws, "C7",  "10-K FY2025 (fiscal year ended December 27, 2025) - SEC filing")
    set_cell(ws, "B8",  "Periodo historico:",       bold=True)
    set_cell(ws, "C8",  "FY2023A - FY2025A (3 anios)")
    set_cell(ws, "B9",  "Periodo proyectado:",      bold=True)
    set_cell(ws, "C9",  "FY2026E - FY2030E (5 anios)")
    set_cell(ws, "B10", "Moneda / unidades:",       bold=True)
    set_cell(ws, "C10", "USD millones (excepto donde se indique)")
    set_cell(ws, "B11", "Convenciones de color:",   bold=True)
    set_cell(ws, "C11", "Hardcodes = azul (#0000FF)  |  Links entre sheets = verde (#3C7D22)  |  Formulas = negro")

    set_cell(ws, "B13", "Indice de hojas", bold=True)
    ws["B13"].font = Font(name=FONT_NAME, size=11, bold=True, color="1F4E78")

    indice = [
        ("1.  Assumptions",    "Drivers e hipotesis del modelo"),
        ("2.  Revenue",        "Buildup de ingresos por segmento"),
        ("3.  COGS",           "Buildup de costos por segmento"),
        ("4.  OpEx",           "Buildup de SG&A"),
        ("5.  NWC",            "Capital de trabajo (DSO / DIO / DPO)"),
        ("6.  CapEx_DA",       "CapEx, Depreciacion y Amortizacion - roll forward PPE/Intangibles"),
        ("7.  Debt",           "Debt schedule e interest expense"),
        ("8.  IS",             "Estado de Resultados (P&L)"),
        ("9.  BS",             "Balance General"),
        ("10. CFS",            "Estado de Flujos de Efectivo"),
    ]
    for i, (sheet, desc) in enumerate(indice):
        set_cell(ws, f"B{15+i}", sheet, bold=True)
        set_cell(ws, f"C{15+i}", desc)

    set_cell(ws, "B27", "Notas y supuestos clave", bold=True)
    ws["B27"].font = Font(name=FONT_NAME, size=11, bold=True, color="1F4E78")
    notas = [
        "- Forecast de revenue por segmento: crecimiento anual aplicado al ultimo dato historico (FY25).",
        "- COGS proyectado como % de ventas por segmento (driver del 10-K = adjusted COGS de Note 21).",
        "- SG&A excluye impairments historicos. Se asume crecimiento anual modesto; impairments = 0 forward.",
        "- NWC proyectado con DSO/DIO/DPO; resto de partidas corrientes se mantienen flat.",
        "- CapEx y depreciacion proyectados como % de ventas; amortizacion de intangibles flat (~$250M/anio).",
        "- Debt: se asume refinanciacion del 100% de vencimientos (debt rollover); interes = tasa promedio aplicada al saldo promedio.",
        "- Impuestos: tax rate estatutaria efectiva proyectada ~25% (mix US + foreign).",
        "- Dividendos: se mantiene DPS de $1.60/accion (politica reciente FY23-FY25).",
        "- Cash se obtiene como plug desde el CFS; el Balance debe cerrar (ver fila Check en BS).",
    ]
    for i, n in enumerate(notas):
        set_cell(ws, f"B{28+i}", n, italic=True)

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 90

# ------------------------------------------------------------
# HOJA 2 - ASSUMPTIONS
# ------------------------------------------------------------
# Mapeo de filas (referenciado por todas las demas hojas)
ASM = {
    "title_row": 1,
    "period_row": 3,
    # Revenue growth
    "rev_hdr": 5,
    "na_g":    6,
    "idm_g":   7,
    "em_g":    8,
    # COGS % sales
    "cogs_hdr": 10,
    "na_cogs_pct":  11,
    "idm_cogs_pct": 12,
    "em_cogs_pct":  13,
    # SG&A
    "opex_hdr": 15,
    "sga_growth": 16,
    "sga_pct":    17,
    # NWC days
    "nwc_hdr": 19,
    "dso":     20,
    "dio":     21,
    "dpo":     22,
    # CapEx / D&A
    "capex_hdr": 24,
    "capex_pct": 25,
    "dep_pct":   26,
    "amort_abs": 27,
    # Debt
    "debt_hdr": 29,
    "int_rate": 30,
    "refinance_pct": 31,
    # Tax & dividends
    "other_hdr": 33,
    "tax_rate":  34,
    "other_exp_pct_sales": 35,
    "dps":       36,
    "shares":    37,
    "share_repch": 38,
}

def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, ASM["period_row"] + 1)

    set_cell(ws, f"A{ASM['title_row']}", "ASSUMPTIONS - Drivers del modelo", bold=True)
    ws.cell(row=ASM['title_row'], column=1).font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")

    period_headers(ws, ASM["period_row"])

    # =========== Revenue growth ===========
    section_header(ws, ASM["rev_hdr"], "1. Revenue growth (YoY %)")
    set_label(ws, ASM["na_g"],  "North America - sales growth")
    set_label(ws, ASM["idm_g"], "International Developed Markets - sales growth")
    set_label(ws, ASM["em_g"],  "Emerging Markets - sales growth")

    # Historicos: formula (current/prior - 1), FY23 queda vacio (no hay FY22)
    for i in range(1, N_HIST):  # i=1,2 -> cols C,D
        col_now  = HIST_COLS[i]
        col_prev = HIST_COLS[i-1]
        set_cell(ws, f"{col_now}{ASM['na_g']}",  f"=Revenue!{col_now}6/Revenue!{col_prev}6-1",  kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{col_now}{ASM['idm_g']}", f"=Revenue!{col_now}7/Revenue!{col_prev}7-1",  kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{col_now}{ASM['em_g']}",  f"=Revenue!{col_now}8/Revenue!{col_prev}8-1",  kind="link", fmt=PCT_FMT)

    # Forecast: hardcodes de crecimiento (calibrados sobre los ultimos anos)
    na_g_fcst  = [-0.020, 0.005, 0.015, 0.020, 0.025]   # NA: declinante pero recuperando
    idm_g_fcst = [0.005,  0.015, 0.020, 0.025, 0.025]
    em_g_fcst  = [0.030,  0.040, 0.045, 0.050, 0.050]
    write_row_values(ws, ASM["na_g"],  na_g_fcst,  start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["idm_g"], idm_g_fcst, start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["em_g"],  em_g_fcst,  start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)

    # =========== COGS % sales ===========
    section_header(ws, ASM["cogs_hdr"], "2. COGS as % of Sales (by segment)")
    set_label(ws, ASM["na_cogs_pct"],  "North America - COGS % sales")
    set_label(ws, ASM["idm_cogs_pct"], "International Developed Markets - COGS % sales")
    set_label(ws, ASM["em_cogs_pct"],  "Emerging Markets - COGS % sales")

    # Historicos: linked desde Revenue & COGS
    for i in range(N_HIST):
        c = HIST_COLS[i]
        set_cell(ws, f"{c}{ASM['na_cogs_pct']}",  f"=COGS!{c}6/Revenue!{c}6",  kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{ASM['idm_cogs_pct']}", f"=COGS!{c}7/Revenue!{c}7", kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{ASM['em_cogs_pct']}",  f"=COGS!{c}8/Revenue!{c}8", kind="link", fmt=PCT_FMT)

    # Forecast: % held at recent average (calibrado al historico FY25)
    na_cogs_fcst  = [0.650, 0.648, 0.645, 0.643, 0.640]
    idm_cogs_fcst = [0.705, 0.702, 0.700, 0.698, 0.695]
    em_cogs_fcst  = [0.730, 0.728, 0.725, 0.723, 0.720]
    write_row_values(ws, ASM["na_cogs_pct"],  na_cogs_fcst,  start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["idm_cogs_pct"], idm_cogs_fcst, start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["em_cogs_pct"],  em_cogs_fcst,  start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)

    # =========== SG&A ===========
    section_header(ws, ASM["opex_hdr"], "3. OpEx (SG&A excl. impairments)")
    set_label(ws, ASM["sga_growth"], "SG&A growth (YoY %)")
    set_label(ws, ASM["sga_pct"],    "SG&A as % of Sales (reference)")

    for i in range(1, N_HIST):
        c, p = HIST_COLS[i], HIST_COLS[i-1]
        set_cell(ws, f"{c}{ASM['sga_growth']}", f"=OpEx!{c}6/OpEx!{p}6-1", kind="link", fmt=PCT_FMT)
    for i in range(N_HIST):
        c = HIST_COLS[i]
        set_cell(ws, f"{c}{ASM['sga_pct']}", f"=OpEx!{c}6/Revenue!{c}10", kind="link", fmt=PCT_FMT)

    sga_g_fcst = [0.015, 0.020, 0.020, 0.025, 0.025]
    write_row_values(ws, ASM["sga_growth"], sga_g_fcst, start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    for i in range(N_FCST):
        c = FCST_COLS[i]
        set_cell(ws, f"{c}{ASM['sga_pct']}", f"=OpEx!{c}6/Revenue!{c}10", kind="link", fmt=PCT_FMT)

    # =========== NWC days ===========
    section_header(ws, ASM["nwc_hdr"], "4. NWC days (DSO / DIO / DPO)")
    set_label(ws, ASM["dso"], "DSO - Days Sales Outstanding")
    set_label(ws, ASM["dio"], "DIO - Days Inventory Outstanding")
    set_label(ws, ASM["dpo"], "DPO - Days Payable Outstanding")

    # Historicos (solo 24 y 25, en BS hay 2 anios) - traemos desde NWC
    for i in range(1, N_HIST):  # FY24 y FY25
        c = HIST_COLS[i]
        set_cell(ws, f"{c}{ASM['dso']}", f"=NWC!{c}11", kind="link", fmt=DAYS_FMT)
        set_cell(ws, f"{c}{ASM['dio']}", f"=NWC!{c}12", kind="link", fmt=DAYS_FMT)
        set_cell(ws, f"{c}{ASM['dpo']}", f"=NWC!{c}13", kind="link", fmt=DAYS_FMT)

    # Forecast (calibrado al ultimo historico FY25)
    write_row_values(ws, ASM["dso"], [33, 33, 33, 33, 33], start_col=FIRST_DATA_COL+N_HIST, fmt=DAYS_FMT)
    write_row_values(ws, ASM["dio"], [69, 68, 68, 67, 67], start_col=FIRST_DATA_COL+N_HIST, fmt=DAYS_FMT)
    write_row_values(ws, ASM["dpo"], [95, 95, 95, 95, 95], start_col=FIRST_DATA_COL+N_HIST, fmt=DAYS_FMT)

    # =========== CapEx & D&A ===========
    section_header(ws, ASM["capex_hdr"], "5. CapEx & D&A")
    set_label(ws, ASM["capex_pct"], "CapEx % of Sales")
    set_label(ws, ASM["dep_pct"],   "Depreciation % of Sales")
    set_label(ws, ASM["amort_abs"], "Amortization of intangibles ($M)")

    for i in range(N_HIST):
        c = HIST_COLS[i]
        set_cell(ws, f"{c}{ASM['capex_pct']}", f"=CapEx_DA!{c}6/Revenue!{c}10", kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{ASM['dep_pct']}",   f"=CapEx_DA!{c}13/Revenue!{c}10", kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{ASM['amort_abs']}", f"=CapEx_DA!{c}14", kind="link", fmt=NUM_FMT)

    write_row_values(ws, ASM["capex_pct"], [0.034, 0.034, 0.033, 0.033, 0.032], start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["dep_pct"],   [0.029, 0.029, 0.028, 0.028, 0.027], start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["amort_abs"], [250, 250, 250, 250, 250],          start_col=FIRST_DATA_COL+N_HIST, fmt=NUM_FMT)

    # =========== Debt ===========
    section_header(ws, ASM["debt_hdr"], "6. Debt")
    set_label(ws, ASM["int_rate"],      "Average interest rate on debt")
    set_label(ws, ASM["refinance_pct"], "Refinancing % of scheduled maturities")

    # Historical interest rate: int expense / avg debt
    set_cell(ws, f"D{ASM['int_rate']}", f"=IS!D14/AVERAGE(BS!C25+BS!C30,BS!D25+BS!D30)", kind="link", fmt=PCT_FMT)
    write_row_values(ws, ASM["int_rate"],      [0.046, 0.046, 0.046, 0.046, 0.046], start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["refinance_pct"], [1.00,  1.00,  1.00,  1.00,  1.00 ], start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)

    # =========== Other ===========
    section_header(ws, ASM["other_hdr"], "7. Tax, Other & Equity")
    set_label(ws, ASM["tax_rate"],            "Effective tax rate")
    set_label(ws, ASM["other_exp_pct_sales"], "Other (income)/expense ($M, flat)")
    set_label(ws, ASM["dps"],                 "Dividends per share ($/share)")
    set_label(ws, ASM["shares"],              "Diluted shares outstanding (M)")
    set_label(ws, ASM["share_repch"],         "Share repurchases ($M)")

    write_row_values(ws, ASM["tax_rate"],            [0.25, 0.25, 0.25, 0.25, 0.25],  start_col=FIRST_DATA_COL+N_HIST, fmt=PCT_FMT)
    write_row_values(ws, ASM["other_exp_pct_sales"], [-100, -100, -100, -100, -100], start_col=FIRST_DATA_COL+N_HIST, fmt=NUM_FMT)
    write_row_values(ws, ASM["dps"],                 [1.60, 1.60, 1.60, 1.60, 1.60], start_col=FIRST_DATA_COL+N_HIST, fmt='$0.00')
    write_row_values(ws, ASM["shares"],              [1184, 1184, 1184, 1184, 1184], start_col=FIRST_DATA_COL+N_HIST, fmt=NUM_FMT)
    write_row_values(ws, ASM["share_repch"],         [500, 500, 500, 500, 500],      start_col=FIRST_DATA_COL+N_HIST, fmt=NUM_FMT)

    # Historicos para tax rate y dividendos: calculados desde IS y CFS
    for i in range(N_HIST):
        c = HIST_COLS[i]
        set_cell(ws, f"{c}{ASM['tax_rate']}", f"=IF(IS!{c}15<>0,IS!{c}16/IS!{c}15,0)", kind="link", fmt=PCT_FMT)

    apply_default_font(ws, max_row=ASM["share_repch"]+3, max_col=10)


# ------------------------------------------------------------
# HOJA 3 - REVENUE
# ------------------------------------------------------------
REV = {
    "title": 1, "period": 3, "hdr_seg": 5,
    "na": 6, "idm": 7, "em": 8,
    "total_hdr": 10,
    "growth_hdr": 12, "na_g": 13, "idm_g": 14, "em_g": 15, "tot_g": 16,
}

def build_revenue(wb):
    ws = wb.create_sheet("Revenue")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, REV["period"] + 1)

    set_cell(ws, "A1", "REVENUE BUILDUP - Net Sales by Segment", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, REV["period"])

    section_header(ws, REV["hdr_seg"], "Net sales by segment ($M)")
    set_label(ws, REV["na"],  "North America")
    set_label(ws, REV["idm"], "International Developed Markets")
    set_label(ws, REV["em"],  "Emerging Markets")

    # --- Historicos: hardcodes en azul ---
    write_row_values(ws, REV["na"],  NA_SALES_HIST,  start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, REV["idm"], IDM_SALES_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, REV["em"],  EM_SALES_HIST,  start_col=FIRST_DATA_COL, fmt=NUM_FMT)

    # --- Forecast: prior * (1 + growth de Assumptions) ---
    for i, c in enumerate(FCST_COLS):
        prev = FCST_COLS[i-1] if i > 0 else LAST_HIST_COL
        set_cell(ws, f"{c}{REV['na']}",  f"={prev}{REV['na']}*(1+Assumptions!{c}{ASM['na_g']})",  kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{REV['idm']}", f"={prev}{REV['idm']}*(1+Assumptions!{c}{ASM['idm_g']})", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{REV['em']}",  f"={prev}{REV['em']}*(1+Assumptions!{c}{ASM['em_g']})",  kind="link", fmt=NUM_FMT)

    # --- Total Net Sales ---
    set_label(ws, REV["total_hdr"], "Total Net Sales", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{REV['total_hdr']}", f"=SUM({c}{REV['na']}:{c}{REV['em']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- Growth rates check ---
    section_header(ws, REV["growth_hdr"], "Growth rates YoY (check)")
    set_label(ws, REV["na_g"],  "North America")
    set_label(ws, REV["idm_g"], "International Developed Markets")
    set_label(ws, REV["em_g"],  "Emerging Markets")
    set_label(ws, REV["tot_g"], "Total Kraft Heinz", bold=True)

    for i in range(1, N_ALL):
        c, p = ALL_COLS[i], ALL_COLS[i-1]
        set_cell(ws, f"{c}{REV['na_g']}",  f"={c}{REV['na']}/{p}{REV['na']}-1",   fmt=PCT_FMT)
        set_cell(ws, f"{c}{REV['idm_g']}", f"={c}{REV['idm']}/{p}{REV['idm']}-1", fmt=PCT_FMT)
        set_cell(ws, f"{c}{REV['em_g']}",  f"={c}{REV['em']}/{p}{REV['em']}-1",   fmt=PCT_FMT)
        set_cell(ws, f"{c}{REV['tot_g']}", f"={c}{REV['total_hdr']}/{p}{REV['total_hdr']}-1", fmt=PCT_FMT, bold=True)

    apply_default_font(ws, max_row=20, max_col=10)


# ------------------------------------------------------------
# HOJA 4 - COGS
# ------------------------------------------------------------
COGS_R = {
    "title": 1, "period": 3, "hdr": 5,
    "na": 6, "idm": 7, "em": 8, "total": 10,
    "pct_hdr": 12, "na_pct": 13, "idm_pct": 14, "em_pct": 15, "gp_pct": 17,
}

def build_cogs(wb):
    ws = wb.create_sheet("COGS")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, COGS_R["period"] + 1)

    set_cell(ws, "A1", "COGS BUILDUP - by Segment (% of Sales)", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, COGS_R["period"])

    section_header(ws, COGS_R["hdr"], "COGS by segment ($M)")
    set_label(ws, COGS_R["na"],  "North America")
    set_label(ws, COGS_R["idm"], "International Developed Markets")
    set_label(ws, COGS_R["em"],  "Emerging Markets")

    write_row_values(ws, COGS_R["na"],  NA_COGS_HIST,  start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, COGS_R["idm"], IDM_COGS_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, COGS_R["em"],  EM_COGS_HIST,  start_col=FIRST_DATA_COL, fmt=NUM_FMT)

    # Forecast: COGS = Revenue_segment * COGS% from Assumptions
    for c in FCST_COLS:
        set_cell(ws, f"{c}{COGS_R['na']}",  f"=Revenue!{c}{REV['na']}*Assumptions!{c}{ASM['na_cogs_pct']}",  kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{COGS_R['idm']}", f"=Revenue!{c}{REV['idm']}*Assumptions!{c}{ASM['idm_cogs_pct']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{COGS_R['em']}",  f"=Revenue!{c}{REV['em']}*Assumptions!{c}{ASM['em_cogs_pct']}",  kind="link", fmt=NUM_FMT)

    set_label(ws, COGS_R["total"], "Total COGS", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{COGS_R['total']}", f"=SUM({c}{COGS_R['na']}:{c}{COGS_R['em']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- COGS % Sales (check) ---
    section_header(ws, COGS_R["pct_hdr"], "COGS as % of Segment Sales (check)")
    set_label(ws, COGS_R["na_pct"],  "North America")
    set_label(ws, COGS_R["idm_pct"], "International Developed Markets")
    set_label(ws, COGS_R["em_pct"],  "Emerging Markets")
    set_label(ws, COGS_R["gp_pct"],  "Gross margin %  (consolidated)", bold=True)

    for c in ALL_COLS:
        set_cell(ws, f"{c}{COGS_R['na_pct']}",  f"={c}{COGS_R['na']}/Revenue!{c}{REV['na']}",  kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{COGS_R['idm_pct']}", f"={c}{COGS_R['idm']}/Revenue!{c}{REV['idm']}", kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{COGS_R['em_pct']}",  f"={c}{COGS_R['em']}/Revenue!{c}{REV['em']}",  kind="link", fmt=PCT_FMT)
        set_cell(ws, f"{c}{COGS_R['gp_pct']}",  f"=1-{c}{COGS_R['total']}/Revenue!{c}{REV['total_hdr']}", kind="link", fmt=PCT_FMT, bold=True)

    apply_default_font(ws, max_row=20, max_col=10)


# ------------------------------------------------------------
# HOJA 5 - OPEX (SG&A)
# ------------------------------------------------------------
OPEX_R = {"title": 1, "period": 3, "hdr": 5, "sga": 6, "growth": 8, "pct": 9}

def build_opex(wb):
    ws = wb.create_sheet("OpEx")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, OPEX_R["period"] + 1)

    set_cell(ws, "A1", "OPEX BUILDUP - SG&A (excluding impairments)", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, OPEX_R["period"])

    section_header(ws, OPEX_R["hdr"], "SG&A excluding impairment losses ($M)")
    set_label(ws, OPEX_R["sga"], "SG&A (excl. impairments)", bold=True)

    # Historicos hardcode
    write_row_values(ws, OPEX_R["sga"], SGA_EX_IMPAIR_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    # Forecast: prior * (1 + growth)
    for i, c in enumerate(FCST_COLS):
        prev = FCST_COLS[i-1] if i > 0 else LAST_HIST_COL
        set_cell(ws, f"{c}{OPEX_R['sga']}", f"={prev}{OPEX_R['sga']}*(1+Assumptions!{c}{ASM['sga_growth']})", kind="link", fmt=NUM_FMT, bold=True)
    for c in HIST_COLS:
        ws[f"{c}{OPEX_R['sga']}"].font = base_font(color=COLOR_HARDCODE, bold=True)
    ws[f"A{OPEX_R['sga']}"].font = base_font(bold=True)
    for c in ALL_COLS:
        ws[f"{c}{OPEX_R['sga']}"].border = TOP_BOT

    # Growth / pct check
    set_label(ws, OPEX_R["growth"], "SG&A growth YoY")
    set_label(ws, OPEX_R["pct"],    "SG&A as % of Sales")
    for i in range(1, N_ALL):
        c, p = ALL_COLS[i], ALL_COLS[i-1]
        set_cell(ws, f"{c}{OPEX_R['growth']}", f"={c}{OPEX_R['sga']}/{p}{OPEX_R['sga']}-1", fmt=PCT_FMT)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{OPEX_R['pct']}", f"={c}{OPEX_R['sga']}/Revenue!{c}{REV['total_hdr']}", kind="link", fmt=PCT_FMT)

    apply_default_font(ws, max_row=15, max_col=10)


# ------------------------------------------------------------
# HOJA 6 - NWC (Net Working Capital)
# ------------------------------------------------------------
NWC_R = {
    "title": 1, "period": 3,
    "bs_hdr": 5, "ar": 6, "inv": 7, "ap": 8,
    "nwc_total": 9,
    "days_hdr": 10, "dso": 11, "dio": 12, "dpo": 13,
    "chg_hdr": 15, "chg_ar": 16, "chg_inv": 17, "chg_ap": 18, "chg_nwc": 19,
}

def build_nwc(wb):
    ws = wb.create_sheet("NWC")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, NWC_R["period"] + 1)

    set_cell(ws, "A1", "NET WORKING CAPITAL BUILDUP (DSO / DIO / DPO)", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, NWC_R["period"])

    section_header(ws, NWC_R["bs_hdr"], "Working capital balance ($M)")
    set_label(ws, NWC_R["ar"],  "Trade receivables (AR)")
    set_label(ws, NWC_R["inv"], "Inventories")
    set_label(ws, NWC_R["ap"],  "Accounts payable")
    set_label(ws, NWC_R["nwc_total"], "Operating NWC (AR + Inv - AP)", bold=True)

    # Historicos: BS muestra solo FY24 (col C) y FY25 (col D); FY23 (col B) lo dejamos en blanco con "-"
    # AR / Inv / AP en FY24 y FY25
    set_cell(ws, f"C{NWC_R['ar']}",  AR_24,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{NWC_R['ar']}",  AR_25,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{NWC_R['inv']}", INV_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{NWC_R['inv']}", INV_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{NWC_R['ap']}",  AP_24,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{NWC_R['ap']}",  AP_25,  kind="hardcode", fmt=NUM_FMT)

    # Forecast: AR = DSO * Sales / 365 ; Inv = DIO * COGS / 365 ; AP = DPO * COGS / 365
    for c in FCST_COLS:
        set_cell(ws, f"{c}{NWC_R['ar']}",  f"=Assumptions!{c}{ASM['dso']}*Revenue!{c}{REV['total_hdr']}/365", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{NWC_R['inv']}", f"=Assumptions!{c}{ASM['dio']}*COGS!{c}{COGS_R['total']}/365",     kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{NWC_R['ap']}",  f"=Assumptions!{c}{ASM['dpo']}*COGS!{c}{COGS_R['total']}/365",     kind="link", fmt=NUM_FMT)

    # NWC total
    for c in ALL_COLS:
        set_cell(ws, f"{c}{NWC_R['nwc_total']}", f"={c}{NWC_R['ar']}+{c}{NWC_R['inv']}-{c}{NWC_R['ap']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- Implied days ---
    set_label(ws, NWC_R["days_hdr"], "Implied days (check)", bold=True, italic=True)
    set_label(ws, NWC_R["dso"], "DSO = AR / Sales * 365")
    set_label(ws, NWC_R["dio"], "DIO = Inv / COGS * 365")
    set_label(ws, NWC_R["dpo"], "DPO = AP / COGS * 365")

    for c in ALL_COLS:
        # DSO
        set_cell(ws, f"{c}{NWC_R['dso']}", f"=IFERROR({c}{NWC_R['ar']}/Revenue!{c}{REV['total_hdr']}*365,0)", kind="link", fmt=DAYS_FMT)
        set_cell(ws, f"{c}{NWC_R['dio']}", f"=IFERROR({c}{NWC_R['inv']}/COGS!{c}{COGS_R['total']}*365,0)", kind="link", fmt=DAYS_FMT)
        set_cell(ws, f"{c}{NWC_R['dpo']}", f"=IFERROR({c}{NWC_R['ap']}/COGS!{c}{COGS_R['total']}*365,0)",  kind="link", fmt=DAYS_FMT)

    # --- Change in WC (uses cash if positive change in AR/Inv, source if change in AP) ---
    section_header(ws, NWC_R["chg_hdr"], "Change in NWC ($M)  - flows to CFS")
    set_label(ws, NWC_R["chg_ar"],  "(Increase) / decrease in AR")
    set_label(ws, NWC_R["chg_inv"], "(Increase) / decrease in Inventories")
    set_label(ws, NWC_R["chg_ap"],  "Increase / (decrease) in AP")
    set_label(ws, NWC_R["chg_nwc"], "Total change in operating NWC", bold=True)

    for i in range(1, N_ALL):
        c, p = ALL_COLS[i], ALL_COLS[i-1]
        set_cell(ws, f"{c}{NWC_R['chg_ar']}",  f"=-({c}{NWC_R['ar']}-{p}{NWC_R['ar']})",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{NWC_R['chg_inv']}", f"=-({c}{NWC_R['inv']}-{p}{NWC_R['inv']})", fmt=NUM_FMT)
        set_cell(ws, f"{c}{NWC_R['chg_ap']}",  f"={c}{NWC_R['ap']}-{p}{NWC_R['ap']}",      fmt=NUM_FMT)
        set_cell(ws, f"{c}{NWC_R['chg_nwc']}", f"=SUM({c}{NWC_R['chg_ar']}:{c}{NWC_R['chg_ap']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    apply_default_font(ws, max_row=22, max_col=10)


# ------------------------------------------------------------
# HOJA 7 - CapEx & D&A
# ------------------------------------------------------------
CDA = {
    "title": 1, "period": 3,
    "capex_hdr": 5, "na_cx": 6, "idm_cx": 7, "em_cx": 8, "corp_cx": 9, "tot_cx": 10,
    "da_hdr": 12, "dep": 13, "amort": 14, "tot_da": 15,
    "ppe_hdr": 17, "ppe_bop": 18, "ppe_add": 19, "ppe_dep": 20, "ppe_eop": 21,
    "int_hdr": 23, "int_bop": 24, "int_amort": 25, "int_eop": 26,
}

def build_capex_da(wb):
    ws = wb.create_sheet("CapEx_DA")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, CDA["period"] + 1)

    set_cell(ws, "A1", "CAPEX & D&A BUILDUP - Roll-forward of PPE and Intangibles", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, CDA["period"])

    # CapEx
    section_header(ws, CDA["capex_hdr"], "CapEx by segment ($M)")
    set_label(ws, CDA["na_cx"],   "North America")
    set_label(ws, CDA["idm_cx"],  "International Developed Markets")
    set_label(ws, CDA["em_cx"],   "Emerging Markets")
    set_label(ws, CDA["corp_cx"], "General corporate")
    set_label(ws, CDA["tot_cx"],  "Total CapEx", bold=True)

    # Historicos hardcodes
    write_row_values(ws, CDA["na_cx"],   NA_CAPEX_HIST,   start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, CDA["idm_cx"],  IDM_CAPEX_HIST,  start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, CDA["em_cx"],   EM_CAPEX_HIST,   start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, CDA["corp_cx"], CORP_CAPEX_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    # Forecast: total CapEx = % sales (luego prorrateamos a segmentos por mix historico FY25)
    for c in FCST_COLS:
        set_cell(ws, f"{c}{CDA['tot_cx']}", f"=Assumptions!{c}{ASM['capex_pct']}*Revenue!{c}{REV['total_hdr']}", kind="link", fmt=NUM_FMT, bold=True)
    # Forecast por segmento: prorratea por share del total CapEx en FY25
    for c in FCST_COLS:
        set_cell(ws, f"{c}{CDA['na_cx']}",   f"={c}{CDA['tot_cx']}*({LAST_HIST_COL}{CDA['na_cx']}/{LAST_HIST_COL}{CDA['tot_cx']})",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['idm_cx']}",  f"={c}{CDA['tot_cx']}*({LAST_HIST_COL}{CDA['idm_cx']}/{LAST_HIST_COL}{CDA['tot_cx']})",  fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['em_cx']}",   f"={c}{CDA['tot_cx']}*({LAST_HIST_COL}{CDA['em_cx']}/{LAST_HIST_COL}{CDA['tot_cx']})",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['corp_cx']}", f"={c}{CDA['tot_cx']}*({LAST_HIST_COL}{CDA['corp_cx']}/{LAST_HIST_COL}{CDA['tot_cx']})", fmt=NUM_FMT)

    # Total CapEx historico via SUM (sobre-escribe los hardcodes anteriores en col total)
    for c in HIST_COLS:
        set_cell(ws, f"{c}{CDA['tot_cx']}", f"=SUM({c}{CDA['na_cx']}:{c}{CDA['corp_cx']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)
    for c in FCST_COLS:
        ws[f"{c}{CDA['tot_cx']}"].border = TOP_BOT

    # D&A
    section_header(ws, CDA["da_hdr"], "Depreciation & Amortization ($M)")
    set_label(ws, CDA["dep"],   "Depreciation (PPE)")
    set_label(ws, CDA["amort"], "Amortization of intangibles")
    set_label(ws, CDA["tot_da"], "Total D&A", bold=True)

    write_row_values(ws, CDA["dep"],   DEP_HIST,   start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, CDA["amort"], AMORT_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)

    for c in FCST_COLS:
        set_cell(ws, f"{c}{CDA['dep']}",   f"=Assumptions!{c}{ASM['dep_pct']}*Revenue!{c}{REV['total_hdr']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['amort']}", f"=Assumptions!{c}{ASM['amort_abs']}", kind="link", fmt=NUM_FMT)

    for c in ALL_COLS:
        set_cell(ws, f"{c}{CDA['tot_da']}", f"={c}{CDA['dep']}+{c}{CDA['amort']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # PPE roll-forward
    section_header(ws, CDA["ppe_hdr"], "PPE roll-forward ($M)")
    set_label(ws, CDA["ppe_bop"], "Beginning PPE, net")
    set_label(ws, CDA["ppe_add"], "(+) CapEx")
    set_label(ws, CDA["ppe_dep"], "(-) Depreciation")
    set_label(ws, CDA["ppe_eop"], "Ending PPE, net", bold=True)

    # Para historicos solo aplica desde FY24 (col C), porque BS muestra FY24 y FY25
    # FY24: bop = ?(no disclosure), lo dejamos como FY25 BOP = PPE FY24
    set_cell(ws, f"C{CDA['ppe_eop']}", PPE_24, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)
    set_cell(ws, f"D{CDA['ppe_eop']}", PPE_25, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Forecast roll-forward
    for i, c in enumerate(FCST_COLS):
        prev = FCST_COLS[i-1] if i > 0 else LAST_HIST_COL
        set_cell(ws, f"{c}{CDA['ppe_bop']}", f"={prev}{CDA['ppe_eop']}", kind="formula", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['ppe_add']}", f"={c}{CDA['tot_cx']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['ppe_dep']}", f"=-{c}{CDA['dep']}",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['ppe_eop']}", f"=SUM({c}{CDA['ppe_bop']}:{c}{CDA['ppe_dep']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Intangibles roll-forward
    section_header(ws, CDA["int_hdr"], "Intangibles roll-forward ($M)")
    set_label(ws, CDA["int_bop"],   "Beginning Intangibles, net")
    set_label(ws, CDA["int_amort"], "(-) Amortization")
    set_label(ws, CDA["int_eop"],   "Ending Intangibles, net", bold=True)

    set_cell(ws, f"C{CDA['int_eop']}", INTANG_24, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)
    set_cell(ws, f"D{CDA['int_eop']}", INTANG_25, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    for i, c in enumerate(FCST_COLS):
        prev = FCST_COLS[i-1] if i > 0 else LAST_HIST_COL
        set_cell(ws, f"{c}{CDA['int_bop']}",   f"={prev}{CDA['int_eop']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['int_amort']}", f"=-{c}{CDA['amort']}",     fmt=NUM_FMT)
        set_cell(ws, f"{c}{CDA['int_eop']}",   f"={c}{CDA['int_bop']}+{c}{CDA['int_amort']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    apply_default_font(ws, max_row=30, max_col=10)


# ------------------------------------------------------------
# HOJA 8 - DEBT SCHEDULE
# ------------------------------------------------------------
DEBT_R = {
    "title": 1, "period": 3,
    "sched_hdr": 5, "bop": 6, "mat": 7, "iss": 8, "fx": 9, "eop": 10,
    "split_hdr": 12, "cur": 13, "lt": 14,
    "int_hdr": 16, "avg_debt": 17, "int_exp": 18,
}

def build_debt(wb):
    ws = wb.create_sheet("Debt")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, DEBT_R["period"] + 1)

    set_cell(ws, "A1", "DEBT SCHEDULE - Total Debt roll-forward & Interest Expense", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, DEBT_R["period"])

    section_header(ws, DEBT_R["sched_hdr"], "Total debt roll-forward ($M)")
    set_label(ws, DEBT_R["bop"], "Beginning total debt (CurLTD + LTD)")
    set_label(ws, DEBT_R["mat"], "(-) Scheduled maturities (10-K Note 17)")
    set_label(ws, DEBT_R["iss"], "(+) New debt issuance (refinancing)")
    set_label(ws, DEBT_R["fx"],  "(+/-) FX / debt premium amortization (held flat)")
    set_label(ws, DEBT_R["eop"], "Ending total debt", bold=True)

    # FY24 y FY25 historicos (BS total debt = CurLTD + LTD)
    total_debt_24 = CURLTD_24 + LTD_24    # 19869
    total_debt_25 = CURLTD_25 + LTD_25    # 21219
    set_cell(ws, f"C{DEBT_R['eop']}", total_debt_24, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)
    set_cell(ws, f"D{DEBT_R['eop']}", total_debt_25, kind="hardcode", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Forecast: BOP = prior EOP; Mat = schedule hardcoded; Issuance = mat * refi%
    mat_schedule = [DEBT_MAT[2026], DEBT_MAT[2027], DEBT_MAT[2028], DEBT_MAT[2029], DEBT_MAT[2030]]
    for i, c in enumerate(FCST_COLS):
        prev = FCST_COLS[i-1] if i > 0 else LAST_HIST_COL
        set_cell(ws, f"{c}{DEBT_R['bop']}", f"={prev}{DEBT_R['eop']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['mat']}", -mat_schedule[i], kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['iss']}", f"=-{c}{DEBT_R['mat']}*Assumptions!{c}{ASM['refinance_pct']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['fx']}",  0, kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['eop']}", f"=SUM({c}{DEBT_R['bop']}:{c}{DEBT_R['fx']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Current vs LT split
    section_header(ws, DEBT_R["split_hdr"], "Current vs Long-term split ($M)")
    set_label(ws, DEBT_R["cur"], "Current portion of LT debt")
    set_label(ws, DEBT_R["lt"],  "Long-term debt")

    set_cell(ws, f"C{DEBT_R['cur']}", CURLTD_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{DEBT_R['cur']}", CURLTD_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{DEBT_R['lt']}",  LTD_24,    kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{DEBT_R['lt']}",  LTD_25,    kind="hardcode", fmt=NUM_FMT)

    # Forecast: current = el monto que vence el ANIO SIGUIENTE; LT = EOP - current
    next_year_mat = mat_schedule[1:] + [mat_schedule[-1]]  # FY26 current = FY27 mat; etc.
    for i, c in enumerate(FCST_COLS):
        set_cell(ws, f"{c}{DEBT_R['cur']}", next_year_mat[i], kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['lt']}",  f"={c}{DEBT_R['eop']}-{c}{DEBT_R['cur']}", fmt=NUM_FMT)

    # Interest expense
    section_header(ws, DEBT_R["int_hdr"], "Interest expense ($M)")
    set_label(ws, DEBT_R["avg_debt"], "Average total debt")
    set_label(ws, DEBT_R["int_exp"],  "Interest expense", bold=True)

    # Historicos: hardcode
    write_row_values(ws, DEBT_R["int_exp"], INT_EXP_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    for c in HIST_COLS:
        ws[f"{c}{DEBT_R['int_exp']}"].font = base_font(color=COLOR_HARDCODE, bold=True)
        ws[f"{c}{DEBT_R['int_exp']}"].border = TOP_BOT

    # avg debt: solo desde FY25 podemos calcularlo (necesita FY24 BOP)
    set_cell(ws, f"D{DEBT_R['avg_debt']}", f"=AVERAGE(C{DEBT_R['eop']},D{DEBT_R['eop']})", fmt=NUM_FMT)
    for i, c in enumerate(FCST_COLS):
        set_cell(ws, f"{c}{DEBT_R['avg_debt']}", f"=AVERAGE({c}{DEBT_R['bop']},{c}{DEBT_R['eop']})", fmt=NUM_FMT)
        set_cell(ws, f"{c}{DEBT_R['int_exp']}",  f"={c}{DEBT_R['avg_debt']}*Assumptions!{c}{ASM['int_rate']}", kind="link", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    apply_default_font(ws, max_row=22, max_col=10)


# ------------------------------------------------------------
# HOJA 9 - INCOME STATEMENT (IS)
# ------------------------------------------------------------
IS_R = {
    "title": 1, "period": 3, "hdr": 5,
    "sales": 6, "cogs": 7, "gp": 8,
    "sga": 9, "gw_imp": 10, "int_imp": 11,
    "op_inc": 12,
    "int_exp": 14, "other": 15,
    "pretax": 16, "tax": 17, "ni": 18,
    "margins_hdr": 20, "gm": 21, "om": 22, "nm": 23, "eps": 24,
}

def build_is(wb):
    ws = wb.create_sheet("IS")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, IS_R["period"] + 1)

    set_cell(ws, "A1", "INCOME STATEMENT (P&L) - Consolidated", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, IS_R["period"])

    section_header(ws, IS_R["hdr"], "Consolidated Income Statement ($M)")

    # --- Sales (link from Revenue) ---
    set_label(ws, IS_R["sales"], "Net sales", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['sales']}", f"=Revenue!{c}{REV['total_hdr']}", kind="link", fmt=NUM_FMT, bold=True)

    # --- COGS (link from COGS) ---
    set_label(ws, IS_R["cogs"], "Cost of products sold")
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['cogs']}", f"=-COGS!{c}{COGS_R['total']}", kind="link", fmt=NUM_FMT)

    # --- Gross Profit ---
    set_label(ws, IS_R["gp"], "Gross profit", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['gp']}", f"={c}{IS_R['sales']}+{c}{IS_R['cogs']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- SG&A (negative, link from OpEx) ---
    set_label(ws, IS_R["sga"], "Selling, general & administrative expense")
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['sga']}", f"=-OpEx!{c}{OPEX_R['sga']}", kind="link", fmt=NUM_FMT)

    # --- Impairments (historicos hardcode, 0 forward) ---
    set_label(ws, IS_R["gw_imp"],  "Goodwill impairment losses")
    set_label(ws, IS_R["int_imp"], "Intangible asset impairment losses")
    write_row_values(ws, IS_R["gw_imp"],  [-x for x in GW_IMPAIR_HIST],  start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, IS_R["int_imp"], [-x for x in INT_IMPAIR_HIST], start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    for c in FCST_COLS:
        set_cell(ws, f"{c}{IS_R['gw_imp']}",  0, kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"{c}{IS_R['int_imp']}", 0, kind="hardcode", fmt=NUM_FMT)

    # --- Operating Income ---
    set_label(ws, IS_R["op_inc"], "Operating income / (loss)", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['op_inc']}", f"={c}{IS_R['gp']}+{c}{IS_R['sga']}+{c}{IS_R['gw_imp']}+{c}{IS_R['int_imp']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- Interest & Other ---
    set_label(ws, IS_R["int_exp"], "Interest expense")
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['int_exp']}", f"=-Debt!{c}{DEBT_R['int_exp']}", kind="link", fmt=NUM_FMT)

    set_label(ws, IS_R["other"], "Other (expense) / income")
    # Historicos: signo invertido (other expense en 10-K es expense positivo, queremos como income positivo aqui)
    # 10-K: 2023=27 (exp), 2024=-85 (inc), 2025=-171 (inc)
    # En IS: other income aumenta utilidad, asi que mostramos como negativo de other expense
    write_row_values(ws, IS_R["other"], [-x for x in OTHER_EXP_HIST], start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    for c in FCST_COLS:
        set_cell(ws, f"{c}{IS_R['other']}", f"=-Assumptions!{c}{ASM['other_exp_pct_sales']}", kind="link", fmt=NUM_FMT)

    # --- Pretax Income ---
    set_label(ws, IS_R["pretax"], "Income / (loss) before income taxes", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['pretax']}", f"={c}{IS_R['op_inc']}+{c}{IS_R['int_exp']}+{c}{IS_R['other']}", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # --- Tax ---
    set_label(ws, IS_R["tax"], "Provision for / (benefit from) income taxes")
    # Historicos hardcodes
    write_row_values(ws, IS_R["tax"], TAX_PROV_HIST, start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    for c in FCST_COLS:
        set_cell(ws, f"{c}{IS_R['tax']}", f"={c}{IS_R['pretax']}*Assumptions!{c}{ASM['tax_rate']}", kind="link", fmt=NUM_FMT)

    # --- Net Income ---
    set_label(ws, IS_R["ni"], "Net income / (loss)", bold=True)
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['ni']}", f"={c}{IS_R['pretax']}-{c}{IS_R['tax']}", fmt=NUM_FMT, bold=True, border=DBL_BOT)

    # --- Margins ---
    section_header(ws, IS_R["margins_hdr"], "Margins & EPS (check)")
    set_label(ws, IS_R["gm"], "Gross margin %")
    set_label(ws, IS_R["om"], "Operating margin %")
    set_label(ws, IS_R["nm"], "Net margin %")
    set_label(ws, IS_R["eps"], "EPS - basic ($)")
    for c in ALL_COLS:
        set_cell(ws, f"{c}{IS_R['gm']}", f"={c}{IS_R['gp']}/{c}{IS_R['sales']}",       fmt=PCT_FMT)
        set_cell(ws, f"{c}{IS_R['om']}", f"={c}{IS_R['op_inc']}/{c}{IS_R['sales']}",  fmt=PCT_FMT)
        set_cell(ws, f"{c}{IS_R['nm']}", f"={c}{IS_R['ni']}/{c}{IS_R['sales']}",      fmt=PCT_FMT)
    # EPS - solo forecast (historicos hardcode)
    eps_hist = [2.33, 2.27, -4.93]
    for i, c in enumerate(HIST_COLS):
        set_cell(ws, f"{c}{IS_R['eps']}", eps_hist[i], kind="hardcode", fmt='$0.00')
    for c in FCST_COLS:
        set_cell(ws, f"{c}{IS_R['eps']}", f"={c}{IS_R['ni']}/Assumptions!{c}{ASM['shares']}", kind="link", fmt='$0.00')

    apply_default_font(ws, max_row=28, max_col=10)


# ------------------------------------------------------------
# HOJA 10 - BALANCE SHEET (BS)
# ------------------------------------------------------------
BS_R = {
    "title": 1, "period": 3,
    "assets_hdr": 5,
    "cash": 6, "ar": 7, "inv": 8, "prepaid": 9, "mktsec": 10, "oth_ca": 11, "ahfs": 12,
    "tot_ca": 13,
    "ppe": 14, "gw": 15, "intang": 16, "oth_nca": 17,
    "tot_assets": 18,
    "liab_hdr": 20,
    "curltd": 21, "ap": 22, "accmkt": 23, "intpay": 24, "oth_cl": 25, "lhfs": 26,
    "tot_cl": 27,
    "ltd": 28, "deftax": 29, "accpost": 30, "ltdefinc": 31, "oth_ncl": 32,
    "tot_liab": 33,
    "eq_hdr": 35,
    "rednci": 36,
    "cstk": 37, "apic": 38, "re": 39, "aoci": 40, "treas": 41, "nci": 42,
    "tot_eq": 43,
    "tot_le": 44,
    "check": 46,
}

def build_bs(wb):
    ws = wb.create_sheet("BS")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, BS_R["period"] + 1)

    set_cell(ws, "A1", "BALANCE SHEET - Consolidated", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, BS_R["period"])

    # ===== ASSETS =====
    section_header(ws, BS_R["assets_hdr"], "ASSETS ($M)")
    labels_assets = [
        (BS_R["cash"],    "Cash and cash equivalents"),
        (BS_R["ar"],      "Trade receivables, net"),
        (BS_R["inv"],     "Inventories"),
        (BS_R["prepaid"], "Prepaid expenses"),
        (BS_R["mktsec"],  "Marketable securities"),
        (BS_R["oth_ca"],  "Other current assets"),
        (BS_R["ahfs"],    "Assets held for sale"),
    ]
    for r, lbl in labels_assets:
        set_label(ws, r, lbl, indent=1)

    # Historicos FY24 y FY25 - hardcodes
    hist_vals_assets = {
        BS_R["cash"]:    (CASH_24, CASH_25),
        BS_R["ar"]:      (AR_24, AR_25),
        BS_R["inv"]:     (INV_24, INV_25),
        BS_R["prepaid"]: (PREPAID_24, PREPAID_25),
        BS_R["mktsec"]:  (MKTSEC_24, MKTSEC_25),
        BS_R["oth_ca"]:  (OTH_CA_24, OTH_CA_25),
        BS_R["ahfs"]:    (AHFS_24, AHFS_25),
    }
    for r, (v24, v25) in hist_vals_assets.items():
        set_cell(ws, f"C{r}", v24, kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"D{r}", v25, kind="hardcode", fmt=NUM_FMT)

    # Forecast - cash es plug desde CFS
    for c in FCST_COLS:
        set_cell(ws, f"{c}{BS_R['cash']}",    f"=CFS!{c}{CFS_R['cash_eop']}",            kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['ar']}",      f"=NWC!{c}{NWC_R['ar']}",                 kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['inv']}",     f"=NWC!{c}{NWC_R['inv']}",                kind="link", fmt=NUM_FMT)
        # Prepaid, mkt securities, oth CA, AHFS = flat
        prev = FCST_COLS[FCST_COLS.index(c)-1] if c != FCST_COLS[0] else LAST_HIST_COL
        set_cell(ws, f"{c}{BS_R['prepaid']}", f"={prev}{BS_R['prepaid']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['mktsec']}",  f"={prev}{BS_R['mktsec']}",  fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['oth_ca']}",  f"={prev}{BS_R['oth_ca']}",  fmt=NUM_FMT)
        # Assets held for sale: en FY26 se vende Italy → 0
        if c == "E":
            set_cell(ws, f"{c}{BS_R['ahfs']}", 0, kind="hardcode", fmt=NUM_FMT)
        else:
            set_cell(ws, f"{c}{BS_R['ahfs']}", f"={prev}{BS_R['ahfs']}", fmt=NUM_FMT)

    # Total current assets
    set_label(ws, BS_R["tot_ca"], "Total current assets", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_ca']}", f"=SUM({c}{BS_R['cash']}:{c}{BS_R['ahfs']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # PPE, GW, Intang, Oth NCA
    set_label(ws, BS_R["ppe"],    "Property, plant and equipment, net", indent=1)
    set_label(ws, BS_R["gw"],     "Goodwill", indent=1)
    set_label(ws, BS_R["intang"], "Intangible assets, net", indent=1)
    set_label(ws, BS_R["oth_nca"],"Other non-current assets", indent=1)

    # Historicos
    set_cell(ws, f"C{BS_R['ppe']}", PPE_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['ppe']}", PPE_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['gw']}",  GW_24,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['gw']}",  GW_25,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['intang']}", INTANG_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['intang']}", INTANG_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['oth_nca']}", OTH_NCA_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['oth_nca']}", OTH_NCA_25, kind="hardcode", fmt=NUM_FMT)

    # Forecast
    for c in FCST_COLS:
        prev = FCST_COLS[FCST_COLS.index(c)-1] if c != FCST_COLS[0] else LAST_HIST_COL
        set_cell(ws, f"{c}{BS_R['ppe']}",    f"=CapEx_DA!{c}{CDA['ppe_eop']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['gw']}",     f"={prev}{BS_R['gw']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['intang']}", f"=CapEx_DA!{c}{CDA['int_eop']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['oth_nca']}", f"={prev}{BS_R['oth_nca']}", fmt=NUM_FMT)

    # Total assets
    set_label(ws, BS_R["tot_assets"], "TOTAL ASSETS", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_assets']}",
                 f"={c}{BS_R['tot_ca']}+{c}{BS_R['ppe']}+{c}{BS_R['gw']}+{c}{BS_R['intang']}+{c}{BS_R['oth_nca']}",
                 fmt=NUM_FMT, bold=True, border=DBL_BOT)

    # ===== LIABILITIES =====
    section_header(ws, BS_R["liab_hdr"], "LIABILITIES ($M)")
    labels_liab = [
        (BS_R["curltd"], "Current portion of long-term debt"),
        (BS_R["ap"],     "Accounts payable"),
        (BS_R["accmkt"], "Accrued marketing"),
        (BS_R["intpay"], "Interest payable"),
        (BS_R["oth_cl"], "Other current liabilities"),
        (BS_R["lhfs"],   "Liabilities held for sale"),
    ]
    for r, lbl in labels_liab:
        set_label(ws, r, lbl, indent=1)

    hist_vals_liab = {
        BS_R["curltd"]: (CURLTD_24, CURLTD_25),
        BS_R["ap"]:     (AP_24, AP_25),
        BS_R["accmkt"]: (ACCMKT_24, ACCMKT_25),
        BS_R["intpay"]: (INTPAY_24, INTPAY_25),
        BS_R["oth_cl"]: (OTH_CL_24, OTH_CL_25),
        BS_R["lhfs"]:   (LHFS_24, LHFS_25),
    }
    for r, (v24, v25) in hist_vals_liab.items():
        set_cell(ws, f"C{r}", v24, kind="hardcode", fmt=NUM_FMT)
        set_cell(ws, f"D{r}", v25, kind="hardcode", fmt=NUM_FMT)

    for c in FCST_COLS:
        prev = FCST_COLS[FCST_COLS.index(c)-1] if c != FCST_COLS[0] else LAST_HIST_COL
        set_cell(ws, f"{c}{BS_R['curltd']}", f"=Debt!{c}{DEBT_R['cur']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['ap']}",     f"=NWC!{c}{NWC_R['ap']}",     kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['accmkt']}", f"={prev}{BS_R['accmkt']}",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['intpay']}", f"={prev}{BS_R['intpay']}",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['oth_cl']}", f"={prev}{BS_R['oth_cl']}",   fmt=NUM_FMT)
        if c == "E":
            set_cell(ws, f"{c}{BS_R['lhfs']}", 0, kind="hardcode", fmt=NUM_FMT)
        else:
            set_cell(ws, f"{c}{BS_R['lhfs']}", f"={prev}{BS_R['lhfs']}", fmt=NUM_FMT)

    set_label(ws, BS_R["tot_cl"], "Total current liabilities", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_cl']}", f"=SUM({c}{BS_R['curltd']}:{c}{BS_R['lhfs']})", fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Non-current liabilities
    set_label(ws, BS_R["ltd"],     "Long-term debt", indent=1)
    set_label(ws, BS_R["deftax"],  "Deferred income taxes", indent=1)
    set_label(ws, BS_R["accpost"], "Accrued postemployment costs", indent=1)
    set_label(ws, BS_R["ltdefinc"],"Long-term deferred income", indent=1)
    set_label(ws, BS_R["oth_ncl"], "Other non-current liabilities", indent=1)

    set_cell(ws, f"C{BS_R['ltd']}", LTD_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['ltd']}", LTD_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['deftax']}",  DEFTAX_24,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['deftax']}",  DEFTAX_25,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['accpost']}", ACCPOST_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['accpost']}", ACCPOST_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['ltdefinc']}",LTDEFINC_24,kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['ltdefinc']}",LTDEFINC_25,kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['oth_ncl']}", OTH_NCL_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['oth_ncl']}", OTH_NCL_25, kind="hardcode", fmt=NUM_FMT)

    for c in FCST_COLS:
        prev = FCST_COLS[FCST_COLS.index(c)-1] if c != FCST_COLS[0] else LAST_HIST_COL
        set_cell(ws, f"{c}{BS_R['ltd']}",     f"=Debt!{c}{DEBT_R['lt']}",      kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['deftax']}",  f"={prev}{BS_R['deftax']}",      fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['accpost']}", f"={prev}{BS_R['accpost']}",     fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['ltdefinc']}",f"={prev}{BS_R['ltdefinc']}",    fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['oth_ncl']}", f"={prev}{BS_R['oth_ncl']}",     fmt=NUM_FMT)

    set_label(ws, BS_R["tot_liab"], "TOTAL LIABILITIES", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_liab']}",
                 f"={c}{BS_R['tot_cl']}+{c}{BS_R['ltd']}+{c}{BS_R['deftax']}+{c}{BS_R['accpost']}+{c}{BS_R['ltdefinc']}+{c}{BS_R['oth_ncl']}",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # ===== EQUITY =====
    section_header(ws, BS_R["eq_hdr"], "EQUITY ($M)")
    set_label(ws, BS_R["rednci"], "Redeemable noncontrolling interest", indent=1)
    set_label(ws, BS_R["cstk"],   "Common stock", indent=1)
    set_label(ws, BS_R["apic"],   "Additional paid-in capital", indent=1)
    set_label(ws, BS_R["re"],     "Retained earnings / (deficit)", indent=1)
    set_label(ws, BS_R["aoci"],   "Accumulated other comprehensive income / (loss)", indent=1)
    set_label(ws, BS_R["treas"],  "Treasury stock, at cost", indent=1)
    set_label(ws, BS_R["nci"],    "Noncontrolling interest", indent=1)

    # Hist
    set_cell(ws, f"C{BS_R['rednci']}", REDNCI_24, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['rednci']}", REDNCI_25, kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['cstk']}",   CSTK_24,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['cstk']}",   CSTK_25,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['apic']}",   APIC_24,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['apic']}",   APIC_25,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['re']}",     RE_24,     kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['re']}",     RE_25,     kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['aoci']}",   AOCI_24,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['aoci']}",   AOCI_25,   kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['treas']}",  TREAS_24,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['treas']}",  TREAS_25,  kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"C{BS_R['nci']}",    NCI_24,    kind="hardcode", fmt=NUM_FMT)
    set_cell(ws, f"D{BS_R['nci']}",    NCI_25,    kind="hardcode", fmt=NUM_FMT)

    # Forecast
    for c in FCST_COLS:
        prev = FCST_COLS[FCST_COLS.index(c)-1] if c != FCST_COLS[0] else LAST_HIST_COL
        set_cell(ws, f"{c}{BS_R['rednci']}", f"={prev}{BS_R['rednci']}", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['cstk']}",   f"={prev}{BS_R['cstk']}",   fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['apic']}",   f"={prev}{BS_R['apic']}",   fmt=NUM_FMT)
        # RE roll: RE_t = RE_{t-1} + NI - dividends
        set_cell(ws, f"{c}{BS_R['re']}",
                 f"={prev}{BS_R['re']}+IS!{c}{IS_R['ni']}-(Assumptions!{c}{ASM['dps']}*Assumptions!{c}{ASM['shares']})",
                 kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['aoci']}",   f"={prev}{BS_R['aoci']}",  fmt=NUM_FMT)
        # Treasury: se reduce por repurchases
        set_cell(ws, f"{c}{BS_R['treas']}",
                 f"={prev}{BS_R['treas']}-Assumptions!{c}{ASM['share_repch']}",
                 kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{BS_R['nci']}", f"={prev}{BS_R['nci']}", fmt=NUM_FMT)

    set_label(ws, BS_R["tot_eq"], "TOTAL EQUITY", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_eq']}",
                 f"=SUM({c}{BS_R['rednci']}:{c}{BS_R['nci']})",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Total L+E
    set_label(ws, BS_R["tot_le"], "TOTAL LIABILITIES AND EQUITY", bold=True)
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{BS_R['tot_le']}", f"={c}{BS_R['tot_liab']}+{c}{BS_R['tot_eq']}", fmt=NUM_FMT, bold=True, border=DBL_BOT)

    # Check
    set_label(ws, BS_R["check"], "CHECK: Total Assets - Total L+E (must = 0)", bold=True, italic=True)
    for c in HIST_COLS + FCST_COLS:
        cell = ws[f"{c}{BS_R['check']}"]
        cell.value = f"={c}{BS_R['tot_assets']}-{c}{BS_R['tot_le']}"
        cell.font = base_font(bold=True)
        cell.number_format = NUM_FMT
        # conditional fill
    # Aplicamos un IF visual via formula condicional simple: la celda muestra 0 si cierra
    # (no aplicamos conditional formatting para mantener el script simple)

    apply_default_font(ws, max_row=BS_R["check"]+2, max_col=10)


# ------------------------------------------------------------
# HOJA 11 - CASH FLOW STATEMENT (CFS)
# ------------------------------------------------------------
CFS_R = {
    "title": 1, "period": 3,
    "cfo_hdr": 5, "ni": 6, "da": 7, "chg_ar": 8, "chg_inv": 9, "chg_ap": 10, "other_cfo": 11,
    "tot_cfo": 12,
    "cfi_hdr": 14, "capex": 15, "other_cfi": 16,
    "tot_cfi": 17,
    "cff_hdr": 19, "debt_iss": 20, "debt_rep": 21, "div": 22, "repch": 23, "other_cff": 24,
    "tot_cff": 25,
    "cash_eop_hdr": 26, "cash_bop": 27, "net_chg": 28, "cash_eop": 29,
}

def build_cfs(wb):
    ws = wb.create_sheet("CFS")
    ws.sheet_view.showGridLines = False
    column_widths(ws)
    freeze_view(ws, CFS_R["period"] + 1)

    set_cell(ws, "A1", "CASH FLOW STATEMENT - Consolidated", bold=True)
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    period_headers(ws, CFS_R["period"])

    # ===== CFO =====
    section_header(ws, CFS_R["cfo_hdr"], "CASH FLOWS FROM OPERATING ACTIVITIES ($M)")
    set_label(ws, CFS_R["ni"],        "Net income / (loss)")
    set_label(ws, CFS_R["da"],        "(+) Depreciation and amortization")
    set_label(ws, CFS_R["chg_ar"],    "(+/-) Change in trade receivables")
    set_label(ws, CFS_R["chg_inv"],   "(+/-) Change in inventories")
    set_label(ws, CFS_R["chg_ap"],    "(+/-) Change in accounts payable")
    set_label(ws, CFS_R["other_cfo"], "(+/-) Other operating items (held flat)")
    set_label(ws, CFS_R["tot_cfo"],   "Net cash from operating activities", bold=True)

    # Forecast (historico se deja en blanco o con valor referencial; los historicos del CFS real estan en el 10-K)
    # Los plugs historicos en "other_cfo" reconcilian con el CFO real del 10-K (3976, 4184, 4462 en FY23/24/25)
    cfo_hist = {
        CFS_R["ni"]:        NI_HIST,                  # [2846, 2746, -5848]
        CFS_R["da"]:        TOT_DA_HIST,              # [961, 948, 968]
        CFS_R["chg_ar"]:    [18, -139, -55],
        CFS_R["chg_inv"]:   [-106, -6, 133],
        CFS_R["chg_ap"]:    [-295, -308, -97],
        # other_cfo: plug = real CFO - (NI + DA + AR + Inv + AP) -> FY23: 552, FY24: 943, FY25: 9361
        CFS_R["other_cfo"]:  [552, 943, 9361],
    }
    for r, vals in cfo_hist.items():
        write_row_values(ws, r, vals, start_col=FIRST_DATA_COL, fmt=NUM_FMT)

    for c in FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['ni']}",      f"=IS!{c}{IS_R['ni']}",                kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['da']}",      f"=CapEx_DA!{c}{CDA['tot_da']}",       kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['chg_ar']}",  f"=NWC!{c}{NWC_R['chg_ar']}",          kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['chg_inv']}", f"=NWC!{c}{NWC_R['chg_inv']}",         kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['chg_ap']}",  f"=NWC!{c}{NWC_R['chg_ap']}",          kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['other_cfo']}", 0, kind="hardcode", fmt=NUM_FMT)

    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['tot_cfo']}",
                 f"=SUM({c}{CFS_R['ni']}:{c}{CFS_R['other_cfo']})",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # ===== CFI =====
    section_header(ws, CFS_R["cfi_hdr"], "CASH FLOWS FROM INVESTING ACTIVITIES ($M)")
    set_label(ws, CFS_R["capex"],     "(-) CapEx")
    set_label(ws, CFS_R["other_cfi"], "(+/-) Other investing (held flat)")
    set_label(ws, CFS_R["tot_cfi"],   "Net cash from investing activities", bold=True)

    # Historicos
    write_row_values(ws, CFS_R["capex"], [-x for x in TOT_CAPEX_HIST], start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    write_row_values(ws, CFS_R["other_cfi"], [97, 1, -1032], start_col=FIRST_DATA_COL, fmt=NUM_FMT)
    # (other_cfi historicos: 2023: 97; 2024: 1 (133-140+8); 2025: -1032 (compra de marketable securities))

    for c in FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['capex']}",     f"=-CapEx_DA!{c}{CDA['tot_cx']}", kind="link", fmt=NUM_FMT)
        # FY26: proceeds netos por Italy Infant Transaction (cerrada Dec 31, 2025 por aprox $146M; neto de LHFS = $144M)
        # 10-K Note 5: Acquisitions and Divestitures
        if c == "E":
            set_cell(ws, f"{c}{CFS_R['other_cfi']}", 144, kind="hardcode", fmt=NUM_FMT,
                     comment="Proceeds netos Italy Infant Transaction (10-K Note 5)")
        else:
            set_cell(ws, f"{c}{CFS_R['other_cfi']}", 0, kind="hardcode", fmt=NUM_FMT)

    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['tot_cfi']}",
                 f"=SUM({c}{CFS_R['capex']}:{c}{CFS_R['other_cfi']})",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # ===== CFF =====
    section_header(ws, CFS_R["cff_hdr"], "CASH FLOWS FROM FINANCING ACTIVITIES ($M)")
    set_label(ws, CFS_R["debt_iss"],  "(+) Proceeds from issuance of long-term debt")
    set_label(ws, CFS_R["debt_rep"],  "(-) Repayments of long-term debt")
    set_label(ws, CFS_R["div"],       "(-) Dividends paid")
    set_label(ws, CFS_R["repch"],     "(-) Share repurchases")
    set_label(ws, CFS_R["other_cff"], "(+/-) Other financing (held flat)")
    set_label(ws, CFS_R["tot_cff"],   "Net cash from financing activities", bold=True)

    # Historicos
    cff_hist = {
        CFS_R["debt_iss"]:  [657, 594, 1620],
        CFS_R["debt_rep"]:  [-848, -618, -678],
        CFS_R["div"]:       [-x for x in DIV_PAID_HIST],
        CFS_R["repch"]:     [-x for x in SHARE_REPCH_HIST],
        CFS_R["other_cff"]: [-67, -65, 141],
    }
    for r, vals in cff_hist.items():
        write_row_values(ws, r, vals, start_col=FIRST_DATA_COL, fmt=NUM_FMT)

    for c in FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['debt_iss']}",  f"=Debt!{c}{DEBT_R['iss']}",        kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['debt_rep']}",  f"=Debt!{c}{DEBT_R['mat']}",        kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['div']}",       f"=-Assumptions!{c}{ASM['dps']}*Assumptions!{c}{ASM['shares']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['repch']}",     f"=-Assumptions!{c}{ASM['share_repch']}", kind="link", fmt=NUM_FMT)
        set_cell(ws, f"{c}{CFS_R['other_cff']}", 0, kind="hardcode", fmt=NUM_FMT)

    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['tot_cff']}",
                 f"=SUM({c}{CFS_R['debt_iss']}:{c}{CFS_R['other_cff']})",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # ===== Cash reconciliation =====
    section_header(ws, CFS_R["cash_eop_hdr"], "CASH RECONCILIATION ($M)")
    set_label(ws, CFS_R["cash_bop"], "Cash, beginning of period")
    set_label(ws, CFS_R["net_chg"],  "Net increase / (decrease) in cash", bold=True)
    set_label(ws, CFS_R["cash_eop"], "Cash, end of period", bold=True)

    # Historicos: BOP cash & equivalents (cash & equivalents only, sin restricted)
    # FY23 BOP ≈ 991 (aprox del 10-K - cash & equivalents al cierre FY22)
    # FY24 BOP = FY23 EOP cash & equivalents (no esta en el BS del 10-K FY25, usamos aprox)
    # FY25 BOP = FY24 EOP cash & equivalents = 1334 (del BS)
    set_cell(ws, f"B{CFS_R['cash_bop']}", 991,    kind="hardcode", fmt=NUM_FMT,
             comment="Cash & equivalents al cierre FY22 - aproximacion (10-K reporta cash + restricted = 1041)")
    set_cell(ws, f"C{CFS_R['cash_bop']}", 1181,   kind="hardcode", fmt=NUM_FMT,
             comment="Cash & equivalents al cierre FY23 - aprox (10-K reporta cash + restricted = 1404)")
    set_cell(ws, f"D{CFS_R['cash_bop']}", CASH_24, kind="hardcode", fmt=NUM_FMT,
             comment="Cash & equivalents al cierre FY24 = 1334 (Balance Sheet del 10-K)")

    # Forecast: FY26 BOP = FY25 BS cash (anclamos al BS para que cierre el modelo)
    # Anios siguientes: BOP = EOP del anio anterior
    for i, c in enumerate(FCST_COLS):
        if i == 0:
            set_cell(ws, f"{c}{CFS_R['cash_bop']}", f"=BS!{LAST_HIST_COL}{BS_R['cash']}", kind="link", fmt=NUM_FMT)
        else:
            prev = FCST_COLS[i-1]
            set_cell(ws, f"{c}{CFS_R['cash_bop']}", f"={prev}{CFS_R['cash_eop']}", fmt=NUM_FMT)

    # Net change in cash
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['net_chg']}",
                 f"={c}{CFS_R['tot_cfo']}+{c}{CFS_R['tot_cfi']}+{c}{CFS_R['tot_cff']}",
                 fmt=NUM_FMT, bold=True, border=TOP_BOT)

    # Cash EOP
    for c in HIST_COLS + FCST_COLS:
        set_cell(ws, f"{c}{CFS_R['cash_eop']}",
                 f"={c}{CFS_R['cash_bop']}+{c}{CFS_R['net_chg']}",
                 fmt=NUM_FMT, bold=True, border=DBL_BOT)

    apply_default_font(ws, max_row=CFS_R["cash_eop"]+2, max_col=10)


# ============================================================
# 6) ENSAMBLAJE
# ============================================================
build_cover(wb)
build_assumptions(wb)
build_revenue(wb)
build_cogs(wb)
build_opex(wb)
build_nwc(wb)
build_capex_da(wb)
build_debt(wb)
build_is(wb)
build_bs(wb)
build_cfs(wb)

# Ordenamos las hojas
wb.move_sheet("Cover", offset=-100)

# ============================================================
# 7) SAVE
# ============================================================
OUT_PATH = "KHC_Financial_Model.xlsx"
wb.save(OUT_PATH)
print(f"Modelo guardado en: {OUT_PATH}")
